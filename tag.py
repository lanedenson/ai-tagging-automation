import openpyxl
from datetime import datetime

# Create log file with timestamp
log_filename = f'data/tag_processing_simple_matching_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
try:
    log_file = open(log_filename, 'w', encoding='utf-8')
except IOError:
    print("Error: Cannot create log file in data directory")
    exit()

def log_message(message):
    """Write message to both console and log file"""
    print(message)
    log_file.write(message + '\n')

# Load the XLSX file
workbook = openpyxl.load_workbook('data/input.xlsx')
worksheet = workbook.active

# Read tags from txt file
try:
    with open('data/tags.txt', 'r', encoding='utf-8') as file:
        tags = [line.strip() for line in file if line.strip()]
except FileNotFoundError:
    log_message("Error: tags.txt file not found in data directory")
    log_file.close()
    exit()

log_message(f"Found {len(tags)} tags to check against.")

# Check if tag columns exist, if not, add them
tag_columns = ['Tag 1', 'Tag 2', 'Tag 3', 'Tag 4', 'Tag 5']
existing_headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]

# Find the last column with content
last_content_column = 1
for col in range(1, worksheet.max_column + 1):
    has_content = False
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=col).value is not None:
            has_content = True
            break
    if has_content:
        last_content_column = col
    else:
        break

log_message(f"Last column with content found at position {last_content_column}")

# Add missing tag column headers
for i, tag_header in enumerate(tag_columns):
    column = None
    # Check if header already exists
    if tag_header in existing_headers:
        column = existing_headers.index(tag_header) + 1
    else:
        # Add new header after the last column
        column = last_content_column + i + 1
        worksheet.cell(row=1, column=column, value=tag_header)
        log_message(f"Added header '{tag_header}' in column {column}")

# Iterate through each row in the XLSX file
for row in range(2, worksheet.max_row + 1):
    # Get the text content from all cells in the row
    row_content = ''
    for col in range(1, last_content_column + 1):
        cell_value = worksheet.cell(row=row, column=col).value
        if cell_value:
            row_content += str(cell_value).lower() + ' '
    
    # Find matching tags for this row
    matching_tags = []
    for tag in tags:
        if tag.lower() in row_content:
            matching_tags.append(tag)
    
    # Add matching tags to the Tag columns
    for i, tag in enumerate(matching_tags[:5]):  # Limit to first 5 matching tags
        column = last_content_column + i + 1
        worksheet.cell(row=row, column=column, value=tag)
        log_message(f"Writing '{tag}' to row {row}, column {column}")

# Save the updated XLSX file
workbook.save('data/output-with-tags.xlsx')

log_message("Processing complete. Results saved in 'output-with-tags-simple-matching.xlsx'")
log_file.close()