import os
from dotenv import load_dotenv
import openpyxl
from datetime import datetime
from anthropic import Anthropic
import re
import unicodedata

# Create data directory if it doesn't exist
os.makedirs('data', exist_ok=True)

# Create log file with timestamp
log_filename = f'output/tag_processing_ai_weighted_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
try:
    log_file = open(log_filename, 'w', encoding='utf-8')
except IOError:
    print("Error: Cannot create log file in data directory")
    exit()

def log_message(message, show_console=False):
    """Write message to log file and optionally to console"""
    if show_console:
        print(message)
    log_file.write(message + '\n')

def clean_text(text):
    """Clean text by removing HTML tags, illegal Excel characters, and normalizing special characters"""
    if not isinstance(text, str):
        return str(text) if text is not None else ""
    
    # Remove control characters and other problematic characters
    text = "".join(char for char in text if ord(char) >= 32 or char in '\n\r\t')
    
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', ' ', text)
    
    # Replace HTML entities
    text = text.replace('&nbsp;', ' ').replace('&amp;', '&')
    
    # Replace multiple spaces with single space
    text = re.sub(r'\s+', ' ', text)
    
    # Replace newlines and carriage returns
    text = text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    
    # Remove or replace illegal Excel characters
    illegal_chars = ['\x00', '\x0B', '\x0C', '\x0E', '\x0F', '\x1F']
    for char in illegal_chars:
        text = text.replace(char, '')
    
    # Remove any remaining control characters
    text = ''.join(char for char in text if unicodedata.category(char)[0] != 'C')
    
    # Truncate text if it's too long (Excel has a 32,767 character limit)
    if len(text) > 32000:
        text = text[:32000]
    
    return text.strip()

# Initialize Anthropic client
try:
    # Load environment variables from .env file
    load_dotenv()
    
    # Get API key from environment variable
    anthropic_key = os.getenv('ANTHROPIC_API_KEY')
    if not anthropic_key:
        raise ValueError("ANTHROPIC_API_KEY environment variable not found")
        
    client = Anthropic(api_key=anthropic_key)
except Exception as e:
    log_message(f"Error initializing Anthropic client: {str(e)}", show_console=True)
    log_file.close()
    exit()

def analyze_with_ai(content, tags):
    """Use Claude to analyze content and suggest relevant tags"""
    try:
        prompt = f"""Given the following content and list of available tags, return ONLY the most relevant tags that match the content.
        Order them by relevance, most relevant first.
        Respond with just the tags, separated by commas, nothing else.
        
        Content: {content}
        Available tags: {', '.join(tags)}"""
        
        response = client.messages.create(
            model="claude-3-haiku-20240307",
            max_tokens=150,
            temperature=0.3,
            messages=[{
                "role": "user",
                "content": prompt
            }]
        )
        suggested_tags = response.content[0].text.split(',')
        return [tag.strip() for tag in suggested_tags if tag.strip() in tags]
    except Exception as e:
        log_message(f"AI analysis error: {str(e)}")
        return []

# Create output directory if it doesn't exist
os.makedirs('output', exist_ok=True)

# Read tags from txt file
try:
    with open('input/tags.txt', 'r', encoding='utf-8') as file:
        tags = [line.strip() for line in file if line.strip()]
except FileNotFoundError:
    log_message("Error: tags.txt file not found in data directory", show_console=True)
    log_file.close()
    exit()

log_message(f"Found {len(tags)} tags to check against.", show_console=True)

# Create a new workbook for combined output
combined_workbook = openpyxl.Workbook()
combined_worksheet = combined_workbook.active
combined_row = 1

# Process all xlsx and csv files in input directory
input_directory = 'input'
if not os.path.exists(input_directory):
    log_message(f"Error: Input directory '{input_directory}' not found", show_console=True)
    log_file.close()
    exit()

input_files = [f for f in os.listdir(input_directory) if f.endswith(('.xlsx', '.csv'))]
if not input_files:
    log_message("Error: No xlsx or csv files found in input directory", show_console=True)
    log_file.close()
    exit()

# Define tag columns
tag_columns = ['Tag 1', 'Tag 2', 'Tag 3', 'Tag 4', 'Tag 5']

for input_file in input_files:
    log_message(f"\nProcessing file: {input_file}", show_console=True)
    
    # Load the file based on its extension
    file_path = os.path.join(input_directory, input_file)
    if input_file.endswith('.csv'):
        # Create a new workbook and load CSV data
        temp_workbook = openpyxl.Workbook()
        worksheet = temp_workbook.active
        
        import csv
        with open(file_path, 'r', encoding='utf-8') as csvfile:
            csv_reader = csv.reader(csvfile)
            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    cleaned_value = clean_text(value)
                    worksheet.cell(row=row_idx, column=col_idx, value=cleaned_value)
    else:  # xlsx file
        temp_workbook = openpyxl.load_workbook(file_path)
        worksheet = temp_workbook.active

    # Get existing headers and find last content column
    last_content_column = 1
    for col in range(1, worksheet.max_column + 1):
        has_content = False
        for row in range(1, worksheet.max_row + 1):
            if worksheet.cell(row=row, column=col).value is not None:
                has_content = True
                break
        if has_content:
            last_content_column = col
        else:
            break

    log_message(f"Last column with content found at position {last_content_column}")

    # Get headers from first file and add to combined worksheet
    if combined_row == 1:
        for col in range(1, last_content_column + 1):
            header = worksheet.cell(row=1, column=col).value
            combined_worksheet.cell(row=1, column=col, value=header)
        
        # Add tag column headers
        for i, tag_header in enumerate(tag_columns):
            column = last_content_column + i + 1
            combined_worksheet.cell(row=1, column=column, value=tag_header)
        combined_row += 1

    total_rows = worksheet.max_row - 1  # Exclude header row
    log_message(f"Processing {total_rows} rows...", show_console=True)

    # Process each row in current file
    for row in range(2, worksheet.max_row + 1):
        # Copy existing content
        for col in range(1, last_content_column + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            cleaned_value = clean_text(cell_value)
            combined_worksheet.cell(row=combined_row, column=col, value=cleaned_value)
        
        # Get the text content from all cells in the row
        row_content = ''
        for col in range(1, last_content_column + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value:
                row_content += str(cell_value) + ' '
        
        # Clean the row content before processing
        row_content = clean_text(row_content)
        
        # Use both keyword matching and AI analysis
        keyword_matches = set()
        for tag in tags:
            if tag.lower() in row_content.lower():
                keyword_matches.add(tag)
        
        ai_matches = set(analyze_with_ai(row_content, tags))
        
        # Combine and prioritize matches
        weighted_tags = []
        for tag in keyword_matches.union(ai_matches):
            weight = 2 if tag in keyword_matches and tag in ai_matches else 1
            weighted_tags.append((tag, weight))
        
        # Sort by weight and AI suggestion order
        ai_order = {tag: idx for idx, tag in enumerate(ai_matches)}
        weighted_tags.sort(key=lambda x: (-x[1], ai_order.get(x[0], len(ai_matches))))
        
        matching_tags = [tag for tag, _ in weighted_tags]
        
        # Log detailed matching information to file only
        log_message(f"Row {row} - Keyword matches: {keyword_matches}")
        log_message(f"Row {row} - AI suggested tags: {ai_matches}")
        log_message(f"Row {row} - Final weighted order: {matching_tags[:5]}")

        # Add matching tags to the Tag columns
        for i, tag in enumerate(matching_tags[:5]):
            column = last_content_column + i + 1
            combined_worksheet.cell(row=combined_row, column=column, value=tag)
        
        # Show progress every 10 rows
        if (row - 1) % 10 == 0:
            progress = f"Processed {row - 1} of {total_rows} rows..."
            log_message(progress, show_console=True)
        
        combined_row += 1

# Save the combined XLSX file
output_filename = os.path.join('output', 'combined-output-with-tags.xlsx')
combined_workbook.save(output_filename)

log_message(f"\nProcessing complete. Combined results saved in '{output_filename}'", show_console=True)
log_file.close()